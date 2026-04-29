"""Render rule-analysis outputs to an .xlsx workbook for download."""

from __future__ import annotations

import io
import math
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from backend.application.analysis.score_aggregator import overall_score
from backend.application.analysis.step_registry import STEP_TITLES


def _is_period_metric_row(item: Any) -> bool:
    return isinstance(item, dict) and "period" in item and "value" in item


def _write_scalar_block(ws, start_row: int, data: dict[str, Any], skip_keys: set[str]) -> int:
    row = start_row
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for key, val in data.items():
        if key in skip_keys:
            continue
        if isinstance(val, (list, dict)):
            continue
        if val is None:
            continue
        if isinstance(val, float) and math.isnan(val):
            continue
        ws.cell(row=row, column=1, value=key).font = bold
        cell = ws.cell(row=row, column=2, value=val)
        if key.endswith("_reason") and isinstance(val, str):
            cell.alignment = wrap
        row += 1
    return row


def _write_table(ws, start_row: int, title: str, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return start_row
    row = start_row
    ws.cell(row=row, column=1, value=title).font = Font(bold=True)
    row += 1
    headers = list(rows[0].keys())
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1
    for item in rows:
        for col, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=item.get(h))
        row += 1
    return row + 1


def _write_step_sheet(ws, step_num: int, title: str, data: dict[str, Any]) -> None:
    ws.cell(row=1, column=1, value=f"Step {step_num}: {title}").font = Font(bold=True, size=12)
    list_keys = {k for k, v in data.items() if isinstance(v, list)}
    row = _write_scalar_block(ws, 3, data, skip_keys=list_keys)
    for key in sorted(list_keys):
        series = data[key]
        if not isinstance(series, list) or not series:
            continue
        if all(_is_period_metric_row(x) for x in series):
            row = _write_table(ws, row, key, series)
        else:
            row = _write_table(
                ws,
                row,
                key,
                [x if isinstance(x, dict) else {"value": x} for x in series],
            )


def _autosize_columns(ws, max_width: int = 50) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, min(len(str(cell.value)), max_width))
        if max_len > 0:
            ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def export_rule_analysis_workbook(
    stock_code: str,
    stock_name: str,
    step_results: dict[int, dict[str, Any]],
    step_errors: dict[int, str],
    scores: list[int],
    titles: dict[int, str] | None = None,
) -> bytes:
    """Build an in-memory .xlsx and return its raw bytes."""
    titles = titles or dict(STEP_TITLES)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.cell(row=1, column=1, value="Stock code").font = Font(bold=True)
    summary.cell(row=1, column=2, value=stock_code)
    summary.cell(row=2, column=1, value="Stock name").font = Font(bold=True)
    summary.cell(row=2, column=2, value=stock_name)
    summary.cell(row=3, column=1, value="Generated at (UTC)").font = Font(bold=True)
    summary.cell(row=3, column=2, value=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"))
    overall = overall_score(scores)
    summary.cell(row=4, column=1, value="Overall score").font = Font(bold=True)
    summary.cell(row=4, column=2, value=overall)

    row = 6
    summary.cell(row=row, column=1, value="Step").font = Font(bold=True)
    summary.cell(row=row, column=2, value="Title").font = Font(bold=True)
    summary.cell(row=row, column=3, value="Score").font = Font(bold=True)
    summary.cell(row=row, column=4, value="Error").font = Font(bold=True)
    row += 1
    for step_num in sorted(set(list(step_results.keys()) + list(step_errors.keys()))):
        summary.cell(row=row, column=1, value=step_num)
        summary.cell(row=row, column=2, value=titles.get(step_num, ""))
        sd = step_results.get(step_num)
        summary.cell(row=row, column=3, value=sd.get("score") if sd else None)
        summary.cell(row=row, column=4, value=step_errors.get(step_num, ""))
        row += 1

    for step_num in sorted(step_results.keys()):
        title = titles.get(step_num, f"Step{step_num}")
        safe_name = f"Step{step_num}"[:31]
        ws = wb.create_sheet(title=safe_name)
        _write_step_sheet(ws, step_num, title, step_results[step_num])
        _autosize_columns(ws)

    _autosize_columns(summary)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_rule_analysis_to_path(
    path: str | Path,
    stock_code: str,
    stock_name: str,
    step_results: dict[int, dict[str, Any]],
    step_errors: dict[int, str],
    scores: list[int],
    titles: dict[int, str] | None = None,
) -> None:
    data = export_rule_analysis_workbook(
        stock_code, stock_name, step_results, step_errors, scores, titles
    )
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_bytes(data)
